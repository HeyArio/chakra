#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Self-check for the chakra report engine. No test framework needed:

    python3 tests/test_engine.py

Builds tiny synthetic workbooks in a temp dir and checks the scoring
formula, blank-answer handling, band/tier boundaries, the 4-week plan
rules and the HTML report structure. Run it after touching scoring,
build_week_plan or the report builder.
"""
import os, sys, tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import nazarban_service as svc
import server

CHECKS = []
def check(label, cond):
    CHECKS.append((label, bool(cond)))
    print(('  ok   ' if cond else '  FAIL ') + label)

def make_workbook(path, answers, weights):
    """answers: list of 140 (int 1-4 or None). weights: {metric: [140 floats]}."""
    wb = openpyxl.Workbook()
    q = wb.create_sheet('Questions')
    r = wb.create_sheet('Responses')
    del wb['Sheet']
    r.cell(1, 2, 'پاسخ (۱ تا ۴)')
    for i in range(svc.TOTAL_QUESTIONS):
        row = i + 2
        if answers[i] is not None:
            r.cell(row, 2, answers[i])
        for key, letter in svc.WEIGHT_COLS.items():
            col = openpyxl.utils.column_index_from_string(letter)
            q.cell(row, col, weights[key][i])
    wb.save(path)

def main():
    td = tempfile.mkdtemp()
    N = svc.TOTAL_QUESTIONS
    ones = {k: [1] * N for k in svc.WEIGHT_COLS}

    print('level_band boundaries (35 / 55 / 75)')
    check('34.9 -> needs_attention', svc.level_band(34.9)[1] == 'needs_attention')
    check('35   -> low',             svc.level_band(35)[1] == 'low')
    check('54.9 -> low',             svc.level_band(54.9)[1] == 'low')
    check('55   -> moderate',        svc.level_band(55)[1] == 'moderate')
    check('74.9 -> moderate',        svc.level_band(74.9)[1] == 'moderate')
    check('75   -> strength',        svc.level_band(75)[1] == 'strength')
    check('None -> empty',           svc.level_band(None) == ('', ''))

    print('tier_for boundaries (40 / 70)')
    check('39.9 -> low',  svc.tier_for(39.9) == 'low')
    check('40   -> mid',  svc.tier_for(40) == 'mid')
    check('70   -> mid',  svc.tier_for(70) == 'mid')
    check('70.1 -> high', svc.tier_for(70.1) == 'high')
    check('None -> low',  svc.tier_for(None) == 'low')

    print('build_week_plan (client rule: needy <=70 bottom-up, then steady)')
    plan = svc.build_week_plan({'root': 80, 'sacral': 50, 'solar': 90, 'heart': 30,
                                'throat': 75, 'thirdeye': 60, 'crown': 85})
    check('weeks = sacral, heart, thirdeye, root',
          [w['chakra'] for w in plan] == ['sacral', 'heart', 'thirdeye', 'root'])
    check('tiers  = mid, low, mid, high',
          [w['tier'] for w in plan] == ['mid', 'low', 'mid', 'high'])
    check('week numbers 1..4', [w['week'] for w in plan] == [1, 2, 3, 4])
    all_high = svc.build_week_plan({k: 90 for k in svc.CHAKRA_KEYS})
    check('all balanced -> first four bottom-up on maintenance',
          [w['chakra'] for w in all_high] == ['root', 'sacral', 'solar', 'heart']
          and all(w['tier'] == 'high' for w in all_high))

    print('score_workbook: all answers = 4, all weights = 1')
    p1 = os.path.join(td, 'all4.xlsx')
    make_workbook(p1, [4] * N, ones)
    d = svc.score_workbook(p1)
    check('every metric = 100', all(v == 100 for v in d['chakras'].values())
          and all(v == 100 for v in d['indices'].values()))
    check('overall = 100, balance = 100, confidence = 100',
          d['overall_score'] == 100 and d['balance'] == 100 and d['confidence'] == 100)
    check('dominant root -> Builder/سازنده',
          d['dominant'] == 'root' and d['archetype'] == 'Builder'
          and d['archetype_fa'] == 'سازنده')

    print('score_workbook: all answers = 1 -> every metric = 0')
    p2 = os.path.join(td, 'all1.xlsx')
    make_workbook(p2, [1] * N, ones)
    d2 = svc.score_workbook(p2)
    check('every metric = 0', all(v == 0 for v in d2['chakras'].values()))

    print('score_workbook: weighted + blank answers (hand-computed)')
    # root: Q1 w=2 a=3, Q2 w=1 a=blank (ignored), Q3 w=1 a=4
    # -> (3*2 + 4*1 - (2+1)) / (3*(2+1)) * 100 = 7/9*100 = 77.8
    answers = [None] * N
    answers[0], answers[2] = 3, 4
    w = {k: [0] * N for k in svc.WEIGHT_COLS}
    w['root'][0], w['root'][1], w['root'][2] = 2, 1, 1
    p3 = os.path.join(td, 'hand.xlsx')
    make_workbook(p3, answers, w)
    d3 = svc.score_workbook(p3)
    check('root = 77.8', d3['chakras']['root'] == 77.8)
    check('unweighted metrics = None', d3['chakras']['crown'] is None)
    check('answered = 2, confidence = 1.4',
          d3['answered'] == 2 and d3['confidence'] == 1.4)
    check('balance None when a chakra is unscored', d3['balance'] is None)

    print('score_workbook: empty questionnaire -> clear error')
    p4 = os.path.join(td, 'empty.xlsx')
    make_workbook(p4, [None] * N, ones)
    try:
        svc.score_workbook(p4)
        check('raises ValueError', False)
    except ValueError as e:
        check('raises ValueError mentioning Responses', 'Responses' in str(e))

    print('build_html structure')
    fonts = svc._load_fonts()
    html_str = svc.build_html(d, fonts, person_name='<b>سارا</b>', date_str='۱ مهر ۱۴۰۵')
    check('7 pages (2 analysis + 4 weeks + closing)', html_str.count('class="page') == 7)
    check('person name is HTML-escaped',
          '&lt;b&gt;سارا&lt;/b&gt;' in html_str and '<b>سارا</b>' not in html_str)
    check('closing greeting uses the name', '&lt;b&gt;سارا&lt;/b&gt; عزیز،' in html_str)
    check('28-day tracker: 6 practices x 28 days = 168 circles',
          html_str.count('class="trk-day"') == 168)
    check('retest nudge present', 'بعد از ۲۸ روز' in html_str)
    check('wordmark everywhere', html_str.count('شاهراه ثروت') >= 14)
    no_name = svc.build_html(d, fonts, person_name='', date_str='۱ مهر ۱۴۰۵')
    check('no name -> generic greeting', 'همراه عزیز،' in no_name)

    print('server filename sanitizing')
    check('persian name kept', server._safe_name('علی رضایی') == 'علی_رضایی.pdf')
    check('empty -> default',  server._safe_name('') == 'chakra-report.pdf')
    check('path chars stripped', '/' not in server._safe_name('../../x') and
          '.' not in server._safe_name('../../x').replace('.pdf', ''))

    failed = [l for l, ok in CHECKS if not ok]
    print(f'\n{len(CHECKS) - len(failed)}/{len(CHECKS)} checks passed')
    if failed:
        sys.exit(1)

if __name__ == '__main__':
    main()
