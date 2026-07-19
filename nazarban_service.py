#!/usr/bin/env python3
"""Nazarban report service for n8n.
Usage: python3 nazarban_service.py <input_xlsx> <output_pdf> [person_name] [date_str]
"""
import sys, os, re, json, math, html, statistics
from datetime import datetime, timezone, timedelta
from typing import Optional
import openpyxl, warnings
from openpyxl.utils import column_index_from_string
from playwright.sync_api import sync_playwright
from program_content import PROGRAM, TIER_LABEL, CHAKRA_THEME, CHAKRA_FREQ, water_note
warnings.filterwarnings('ignore')

_HERE = os.path.dirname(os.path.abspath(__file__))
TOTAL_QUESTIONS = 140

def _load_fonts() -> dict:
    path = os.path.join(_HERE, 'fonts_b64.json')
    with open(path) as f:
        return json.load(f)

# ===== scoring engine =====

# Column letters in Questions sheet -> metric key (matches client's SUMPRODUCT refs)
WEIGHT_COLS = {
    'root': 'P', 'sacral': 'Q', 'solar': 'R', 'heart': 'S',
    'throat': 'T', 'thirdeye': 'U', 'crown': 'V',
    'financial': 'W', 'emotional': 'X', 'health': 'Y',
    'receptivity': 'Z', 'intuition': 'AA',
}
CHAKRA_KEYS = ['root', 'sacral', 'solar', 'heart', 'throat', 'thirdeye', 'crown']

# Canonical Persian labels — single source of truth for all chakra/index names
CHAKRA_LABEL = {
    'root': 'ریشه', 'sacral': 'خاجی', 'solar': 'خورشیدی', 'heart': 'قلب',
    'throat': 'گلو', 'thirdeye': 'چشم سوم', 'crown': 'تاج',
}
INDEX_LABEL = {
    'financial': 'ثروت', 'emotional': 'عاطفی', 'health': 'سلامتی',
    'receptivity': 'دریافت', 'intuition': 'شهود',  # v2 axis names (intuition kept for legacy)
}
FA = {**CHAKRA_LABEL, **INDEX_LABEL}

# Chakra -> archetype (from the client's Archetypes sheet)
ARCHETYPE = {
    'root': 'Builder', 'sacral': 'Creator', 'solar': 'Leader', 'heart': 'Healer',
    'throat': 'Messenger', 'thirdeye': 'Visionary', 'crown': 'Mystic',
}
ARCHETYPE_FA = {
    'Builder': 'سازنده', 'Creator': 'خالق', 'Leader': 'رهبر', 'Healer': 'شفادهنده',
    'Messenger': 'پیام‌رسان', 'Visionary': 'بینا', 'Mystic': 'عارف',
}

# ===== Jalali date =====

_J_MONTHS = ['فروردین', 'اردیبهشت', 'خرداد', 'تیر', 'مرداد', 'شهریور',
             'مهر', 'آبان', 'آذر', 'دی', 'بهمن', 'اسفند']
_FA_DIGITS = str.maketrans('0123456789', '۰۱۲۳۴۵۶۷۸۹')

def _greg_to_jalali(gy: int, gm: int, gd: int) -> tuple[int, int, int]:
    """Standard civil Gregorian -> Jalali conversion (jdf algorithm)."""
    g_d_m = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    gy2 = gy - 1600
    days = (365 * gy2 + (gy2 + 3) // 4 - (gy2 + 99) // 100 + (gy2 + 399) // 400
            - 80 + gd + g_d_m[gm - 1])
    if gm > 2 and ((gy % 4 == 0 and gy % 100 != 0) or gy % 400 == 0):
        days += 1
    jy = 979 + 33 * (days // 12053); days %= 12053
    jy += 4 * (days // 1461);        days %= 1461
    if days > 365:
        jy += (days - 1) // 365
        days = (days - 1) % 365
    jm = 1 + days // 31 if days < 186 else 7 + (days - 186) // 30
    jd = 1 + (days % 31 if days < 186 else (days - 186) % 30)
    return jy, jm, jd

def today_jalali() -> str:
    """Today's Jalali date in Iran time (UTC+3:30), e.g. '۲۰ تیر ۱۴۰۵'."""
    now = datetime.now(timezone(timedelta(hours=3, minutes=30)))
    jy, jm, jd = _greg_to_jalali(now.year, now.month, now.day)
    return f'{jd} {_J_MONTHS[jm - 1]} {jy}'.translate(_FA_DIGITS)

_FA2EN = str.maketrans('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789')

def jalali_display(raw) -> str:
    """Format a Porsline Jalali datetime string ('1405/04/23-18:51:53', digits
    Latin or Persian) as the report's date style '۲۳ تیر ۱۴۰۵'. '' if unparseable."""
    if not raw:
        return ''
    s = str(raw).translate(_FA2EN)
    m = re.match(r'\s*(\d{3,4})\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})', s)
    if not m:
        return ''
    jy, jm, jd = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not 1 <= jm <= 12:
        return ''
    return f'{jd} {_J_MONTHS[jm - 1]} {jy}'.translate(_FA_DIGITS)


def level_band(score: Optional[float]) -> tuple[str, str]:
    """v2 workbook bands (Calculations!C / Interpretation sheet).
    <40 پرچالش · <60 نیازمند توجه · <78 متعادل نسبی · ≥78 قوی.
    Internal keys kept stable (needs_attention/low/moderate/strength) so the
    report CSS + suggestion copy don't have to change."""
    if score is None:
        return ('', '')
    if score < 40:  return ('پرچالش',       'needs_attention')
    if score < 60:  return ('نیازمند توجه',  'low')
    if score < 78:  return ('متعادل نسبی',   'moderate')
    return ('قوی', 'strength')

# ===== v2 survey scoring (Porsline export + bundled scoring model) =====
# The bot now receives a Porsline survey export (single "Results" sheet)
# that carries ONLY the questionnaire answers (as text) plus the person's
# name and dates. The scoring brain — questions, the four options, 1-4
# scores, reverse flags and per-metric weights — lives in the bundled
# master workbook `scoring_model.xlsx` (the client's own file, the single
# source of truth). We map each text answer back to its 1-4 option and
# apply the workbook's exact formulas:
#     F      = 5-answer if question is معکوس (reverse) else answer
#     metric = SUMPRODUCT(F, weight) / (4 * SUM(weight)) * 100

SCORING_MODEL_FILE = 'scoring_model.xlsx'

# master Questions weight columns J..T  ->  report metric keys
# (financial=Wealth/ثروت, receptivity=Receiving/دریافت — v2 has no intuition)
_MODEL_METRIC_COL = {
    'root': 'J', 'sacral': 'K', 'solar': 'L', 'heart': 'M', 'throat': 'N',
    'thirdeye': 'O', 'crown': 'P',
    'financial': 'Q', 'emotional': 'R', 'health': 'S', 'receptivity': 'T',
}
SURVEY_METRICS = list(_MODEL_METRIC_COL)  # 7 chakras + 4 axes

_ZW = ['‌', '‍', '‎', '‏', '﻿', '\xa0']

def _norm_fa(s) -> str:
    """Normalize Persian text for robust matching: strip ZWNJ/joiners, unify
    ی/ي and ک/ك and alef forms, drop incidental punctuation, collapse spaces."""
    if s is None:
        return ''
    s = str(s)
    for z in _ZW:
        s = s.replace(z, ' ')
    s = (s.replace('ي', 'ی').replace('ك', 'ک').replace('ۀ', 'ه')
           .replace('ة', 'ه').replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا'))
    s = re.sub(r'[؟?.…،,!:;«»"\'()\-]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

_MODEL_CACHE = None

def _load_scoring_model() -> list:
    """Parse the bundled master workbook once. Returns one dict per question:
    normalized question text, normalized options, reverse flag, weights."""
    global _MODEL_CACHE
    if _MODEL_CACHE is not None:
        return _MODEL_CACHE
    path = os.path.join(_HERE, SCORING_MODEL_FILE)
    wb = openpyxl.load_workbook(path, data_only=True)
    q = wb['Questions']
    cols = {k: column_index_from_string(v) for k, v in _MODEL_METRIC_COL.items()}
    out = []
    for row in range(2, q.max_row + 1):
        qtext = q.cell(row, 4).value                      # D سؤال
        if qtext is None or str(qtext).strip() == '':
            continue
        opts = [q.cell(row, c).value for c in (5, 6, 7, 8)]  # E-H گزینه ۱..۴
        stype = q.cell(row, 9).value                          # I نوع امتیاز
        weights = {k: float(q.cell(row, ci).value or 0) for k, ci in cols.items()}
        out.append({
            'q_norm': _norm_fa(qtext),
            'opt_norm': [_norm_fa(o) for o in opts],
            'reverse': 'معکوس' in str(stype or ''),
            'weights': weights,
        })
    if not out:
        raise RuntimeError('scoring_model.xlsx contains no questions')
    _MODEL_CACHE = out
    return out

# Porsline "Results" header keywords used to locate the special columns.
# Columns are found by header TEXT, never by letter: Porsline shifts columns
# when the survey gains a field (the batch export added «شماره موبایل» before
# the dates), and header matching absorbs that.
_P_NAME_KEY = 'نام و نام خانوادگی'
_P_PHONE_KEY = 'موبایل'
_P_SKIP_HDRS = ('پاسخنامه', 'شناسه پاسخ دهنده')

def read_porsline_all(path: str) -> list:
    """Read a Porsline survey export ("Results" sheet). Returns one dict per
    respondent — name, phone, start/end dates and a {normalized-question ->
    answer-text} map — in file order. A file may hold a single respondent
    (the old per-person export) or a whole batch, one row each; blank rows
    (Porsline pads the sheet with them) are skipped wherever they appear."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Results']
    header = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(c.value in (None, '') for c in row):
            continue
        name = phone = start = end = None
        answers = {}
        for c, cell in enumerate(row, 1):
            h = header[c - 1] if c <= len(header) else None
            if h is None or str(h).strip() == '':
                continue
            hs = str(h)
            val = cell.value
            if _P_NAME_KEY in hs:
                name = val
            elif _P_PHONE_KEY in hs:
                phone = val
            elif 'تاریخ شروع' in hs:
                start = val
            elif 'تاریخ اتمام' in hs:
                end = val
            elif any(k in hs for k in _P_SKIP_HDRS):
                continue
            else:
                answers[_norm_fa(hs)] = val
        out.append({
            'name': str(name).strip() if name not in (None, '') else '',
            'phone': str(phone).strip() if phone not in (None, '') else '',
            'start_date': str(start).strip() if start not in (None, '') else '',
            'end_date': str(end).strip() if end not in (None, '') else '',
            'answers': answers,
        })
    if not out:
        raise RuntimeError('Porsline export has no response row')
    return out

def read_porsline(path: str) -> dict:
    """First respondent only — kept for callers that predate batch files."""
    return read_porsline_all(path)[0]

def _match_option(ans_norm: str, opts_norm: list) -> Optional[int]:
    """Return the 0-based option index for an answer text, or None."""
    if not ans_norm:
        return None
    for i, on in enumerate(opts_norm):
        if on == ans_norm:
            return i
    for i, on in enumerate(opts_norm):   # tolerant containment fallback
        if on and (on in ans_norm or ans_norm in on):
            return i
    return None

def _score_respondent(pors: dict) -> dict:
    """Score one parsed respondent (a read_porsline_all() entry) against the
    bundled v2 scoring model and return the data shape build_html() consumes,
    plus person/dates."""
    model = _load_scoring_model()
    ans_by_q = pors['answers']

    total = len(model)
    resolved = []          # (question, F-score 1..4 reverse-adjusted or None)
    unmatched = []
    for m in model:
        idx = _match_option(_norm_fa(ans_by_q.get(m['q_norm'])), m['opt_norm'])
        if idx is None:
            unmatched.append(m['q_norm'])
            resolved.append((m, None))
            continue
        ans = idx + 1
        resolved.append((m, (5 - ans) if m['reverse'] else ans))
    answered = sum(1 for _, f in resolved if f is not None)
    if unmatched:
        print(json.dumps({'warn': 'unmatched_answers', 'count': len(unmatched)},
                         ensure_ascii=False), file=sys.stderr)

    def metric_score(key):
        num = wsum = 0.0
        for m, f in resolved:
            w = m['weights'][key]
            wsum += w
            if f is not None:
                num += f * w
        if wsum == 0:
            return 0.0
        return round(num / (4 * wsum) * 100, 1)

    scores = {k: metric_score(k) for k in SURVEY_METRICS}
    confidence = round(answered / total * 100, 1) if total else 0.0

    chakra_vals = [scores[k] for k in CHAKRA_KEYS]
    balance = round(100 - statistics.pstdev(chakra_vals), 1) if len(chakra_vals) == 7 else None
    dominant = max(CHAKRA_KEYS, key=lambda k: scores[k] if scores[k] is not None else -1)
    arche = ARCHETYPE[dominant]

    detail = {}
    for k, v in scores.items():
        band_fa, band_key = level_band(v)
        detail[k] = {'label_fa': FA[k], 'score': v, 'level_fa': band_fa, 'level': band_key}

    overall = round(sum(chakra_vals) / len(chakra_vals), 1) if chakra_vals else None

    return {
        'metrics': detail,
        'chakras': {k: scores[k] for k in CHAKRA_KEYS},
        'indices': {k: scores[k] for k in INDEX_ORDER},
        'overall_score': overall,
        'confidence': confidence,
        'answered': answered,
        'total_questions': total,
        'balance': balance,
        'dominant': dominant,
        'dominant_fa': FA[dominant],
        'archetype': arche,
        'archetype_fa': ARCHETYPE_FA[arche],
        'person_name': pors['name'],
        'phone': pors.get('phone', ''),
        'start_date': pors['start_date'],
        'end_date': pors['end_date'],
        # report date = survey completion time («تاریخ اتمام»), else start
        'report_date': jalali_display(pors['end_date']) or jalali_display(pors['start_date']),
    }


def score_survey(path: str) -> dict:
    """Score the first respondent of a Porsline export (pre-batch behavior)."""
    return _score_respondent(read_porsline(path))


def score_workbook_all(path: str) -> list:
    """Score every respondent in the workbook, in file order. A Porsline
    export (single 'Results' sheet) may carry one row or a whole batch; a
    legacy Questions/Responses workbook is always a single respondent."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    if 'Results' in sheets:
        return [_score_respondent(p) for p in read_porsline_all(path)]
    return [_score_legacy_workbook(path)]


def score_workbook(path: str) -> dict:
    """First respondent only — kept for callers that predate batch files."""
    return score_workbook_all(path)[0]


def _score_legacy_workbook(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    q = wb['Questions']
    r = wb['Responses']

    # answers: row 2..141 -> int 1..4 (blank allowed)
    answers = []
    for row in range(2, TOTAL_QUESTIONS + 2):
        v = r.cell(row, 2).value
        answers.append(int(v) if v not in (None, '') else None)

    # weights per metric (row-aligned with answers)
    weights = {}
    for key, letter in WEIGHT_COLS.items():
        ci = column_index_from_string(letter)
        weights[key] = [q.cell(row, ci).value or 0 for row in range(2, TOTAL_QUESTIONS + 2)]

    def metric_score(key):
        # client's formula:
        # (SUMPRODUCT(ans,w,answered) - SUMPRODUCT(w,answered)) / (3*SUMPRODUCT(w,answered)) *100
        w = weights[key]
        num = 0.0; wsum = 0.0
        for a, wi in zip(answers, w):
            if a is None or wi == 0:
                continue
            num += a * wi
            wsum += wi
        if wsum == 0:
            return None
        raw = (num - wsum) / (3 * wsum) * 100
        return round(raw, 1)

    scores = {k: metric_score(k) for k in WEIGHT_COLS}

    # completion / confidence
    answered = sum(1 for a in answers if a is not None)
    confidence = round(answered / TOTAL_QUESTIONS * 100, 1)

    # overall balance = 100 - population stdev of 7 chakra scores
    chakra_vals = [scores[k] for k in CHAKRA_KEYS if scores[k] is not None]
    balance = round(100 - statistics.pstdev(chakra_vals), 1) if len(chakra_vals) == 7 else None

    # dominant chakra = max of 7
    dominant = max(CHAKRA_KEYS, key=lambda k: (scores[k] if scores[k] is not None else -1))
    arche = ARCHETYPE[dominant]

    # assemble per-metric detail with bands
    detail = {}
    for k, v in scores.items():
        band_fa, band_key = level_band(v)
        detail[k] = {
            'label_fa': FA[k], 'score': v,
            'level_fa': band_fa, 'level': band_key,
        }

    overall = round(sum(chakra_vals) / len(chakra_vals), 1) if chakra_vals else None

    return {
        'metrics': detail,
        'chakras': {k: scores[k] for k in CHAKRA_KEYS},
        'indices': {k: scores[k] for k in ['financial','emotional','health','receptivity','intuition']},
        'overall_score': overall,
        'confidence': confidence,
        'answered': answered,
        'balance': balance,
        'dominant': dominant,
        'dominant_fa': FA[dominant],
        'archetype': arche,
        'archetype_fa': ARCHETYPE_FA[arche],
    }


# ===== 4-week growth program =====
# Client's rule: walk the chakras bottom-up (root -> crown); any chakra
# scoring 70 or below needs work and claims the next free week (week 1 =
# lowest such chakra). If fewer than 4 need work, remaining weeks are
# filled with the balanced (>70) chakras, still bottom-up, on their
# maintenance program. Which of the 3 prescriptions a week uses depends
# only on that chakra's score: <40 low, 40-70 mid, >70 high.

BOTTOM_UP = ['root', 'sacral', 'solar', 'heart', 'throat', 'thirdeye', 'crown']
PROGRAM_WEEKS = 4

def tier_for(score: Optional[float]) -> str:
    s = score if score is not None else 0
    if s < 40: return 'low'
    if s <= 70: return 'mid'
    return 'high'

def build_week_plan(chakras: dict) -> list:
    needy  = [k for k in BOTTOM_UP if (chakras.get(k) or 0) <= 70]
    steady = [k for k in BOTTOM_UP if k not in needy]
    picked = (needy + steady)[:PROGRAM_WEEKS]
    return [{'week': i, 'chakra': k, 'score': (chakras.get(k) or 0),
             'tier': tier_for(chakras.get(k))}
            for i, k in enumerate(picked, 1)]


# ===== report builder =====

CHAKRA_ORDER = ['crown','thirdeye','throat','heart','solar','sacral','root']  # top->bottom (crown high)
CHAKRA_COLOR = {
    'root':'#E5484D','sacral':'#F2802B','solar':'#F5C542','heart':'#4FB477',
    'throat':'#3DA5D9','thirdeye':'#6A6AE3','crown':'#A65FD9',
}
INDEX_ORDER = ['financial','emotional','health','receptivity']  # v2: 4 axes (no intuition)
INDEX_COLOR = '#8b7fb5'

ARCHE_DESC = {  # from client's Archetypes sheet: positive role + shadow
    'Builder':   ('ثبات‌ساز، قابل اعتماد، عملی', 'ترس از تغییر، کنترل‌گری'),
    'Creator':   ('خلاق، جذب‌کننده فرصت، احساسی', 'پراکندگی، وابستگی به لذت'),
    'Leader':    ('قاطع، هدفمند، اثرگذار', 'سلطه‌گری یا ترس از شکست'),
    'Healer':    ('همدل، پیونددهنده، بخشنده', 'فداکاری افراطی، مرز ضعیف'),
    'Messenger': ('شفاف، مذاکره‌گر، روایت‌ساز', 'سکوت یا تندگویی'),
    'Visionary': ('بینش‌مند، استراتژیک، شهودی', 'خیال‌پردازی بدون اقدام'),
    'Mystic':    ('معنادار، الهام‌بخش، متصل', 'فرار از واقعیت، پراکندگی معنوی'),
}
# Interpretation copy per band (from client's Interpretation sheet)
BAND_SUGGEST = {
    'needs_attention': 'تمرین‌های پایه، ثبت احساسات، کاهش فشار و بررسی ریشه‌های رفتاری.',
    'low':             'تمرین‌های کوچک روزانه و ایجاد عادت‌های قابل سنجش.',
    'moderate':        'تمرکز روی تقویت پیوستگی و تبدیل مهارت به عادت.',
    'strength':        'استفاده آگاهانه از این قوت برای حمایت از حوزه‌های ضعیف‌تر.',
}
BAND_LABEL = {
    'needs_attention':'پرچالش','low':'نیازمند توجه',
    'moderate':'متعادل نسبی','strength':'قوی',
}

def _radar_svg(chakras: dict) -> str:
    """7-axis radar, restyled: dotted rings, gradient fill. Persian labels outside."""
    keys = ['crown','thirdeye','throat','heart','solar','sacral','root']
    cx, cy, R = 210, 205, 150
    n = 7
    def pt(i, r):
        ang = -math.pi/2 + i*2*math.pi/n
        return (cx + r*math.cos(ang), cy + r*math.sin(ang))
    rings = ''
    for frac in (0.25,0.5,0.75,1.0):
        pts = ' '.join(f'{pt(i,R*frac)[0]:.1f},{pt(i,R*frac)[1]:.1f}' for i in range(n))
        rings += f'<polygon points="{pts}" fill="none" stroke="#ffffff" stroke-opacity="0.10" stroke-dasharray="2 4"/>'
    spokes=''; labs=''
    for i,k in enumerate(keys):
        x,y = pt(i,R)
        spokes += f'<line x1="{cx}" y1="{cy}" x2="{x:.1f}" y2="{y:.1f}" stroke="#ffffff" stroke-opacity="0.08"/>'
        lx,ly = pt(i,R+26)
        labs += f'<text x="{lx:.1f}" y="{ly:.1f}" fill="#b9afce" font-size="13" text-anchor="middle" dominant-baseline="middle">{CHAKRA_LABEL[k]}</text>'
    dpts=[]
    for i,k in enumerate(keys):
        v=(chakras.get(k) or 0)/100
        x,y=pt(i,R*v); dpts.append((x,y))
    poly=' '.join(f'{x:.1f},{y:.1f}' for x,y in dpts)
    dots=''.join(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="#cbb8ef"/>' for x,y in dpts)
    return f'''<svg viewBox="0 0 420 430" width="330" xmlns="http://www.w3.org/2000/svg">
      <defs><radialGradient id="rg" cx="50%" cy="50%" r="60%">
        <stop offset="0%" stop-color="#7c5cff" stop-opacity="0.55"/>
        <stop offset="100%" stop-color="#7c5cff" stop-opacity="0.08"/>
      </radialGradient></defs>
      {rings}{spokes}
      <polygon points="{poly}" fill="url(#rg)" stroke="#a78bff" stroke-width="1.5"/>
      {dots}{labs}
    </svg>'''

def _spine(chakras: dict) -> str:
    """Signature element: vertical stack of 7 luminous bars = the energy spine."""
    rows=''
    for k in CHAKRA_ORDER:
        v=chakras.get(k) or 0
        c=CHAKRA_COLOR[k]
        rows+=f'''<div class="spine-row">
          <div class="spine-num">{v:.0f}</div>
          <div class="spine-track"><div class="spine-fill" style="width:{v:.0f}%;background:{c};box-shadow:0 0 10px {c}88"></div></div>
          <div class="spine-dot" style="background:{c};box-shadow:0 0 8px {c}"></div>
          <div class="spine-label">{CHAKRA_LABEL[k]}</div>
        </div>'''
    return f'<div class="spine">{rows}</div>'

def _bars(metrics: dict, keys: list, labelmap: dict) -> str:
    out=''
    for k in keys:
        m=metrics[k]; v=m['score'] or 0
        out+=f'''<div class="bar-row">
          <div class="bar-label">{labelmap[k]}</div>
          <div class="bar-track"><div class="bar-fill" style="width:{v:.0f}%"></div></div>
          <div class="bar-val">{v:.0f}</div>
          <div class="bar-tag tag-{m['level']}">{BAND_LABEL[m['level']]}</div>
        </div>'''
    return out

def _interp_rows(metrics: dict, labelmap: dict) -> str:
    """Build per-chakra interpretation rows HTML."""
    parts = []
    for k in CHAKRA_ORDER:
        m = metrics[k]
        score = m['score'] or 0
        color = CHAKRA_COLOR[k]
        parts.append(
            f'<div class="interp-row" style="border-color:{color}">'
            f'<div class="t"><span>{labelmap[k]}</span>'
            f'<span style="color:{color}">{score:.0f}% · {BAND_LABEL[m["level"]]}</span></div>'
            f'<div class="s">{BAND_SUGGEST[m["level"]]}</div>'
            f'</div>'
        )
    return ''.join(parts)

_FA_DIGIT = {1: '۱', 2: '۲', 3: '۳', 4: '۴'}

# 24x24 stroke line icons for the six prescription sections
_ICON = {
    'sleep':    '<path d="M20.2 14.8A8.5 8.5 0 1 1 9.2 3.8a7.2 7.2 0 0 0 11 11z"/>',
    'water':    '<path d="M12 3.6s6 6.3 6 10.3a6 6 0 0 1-12 0c0-4 6-10.3 6-10.3z"/>',
    'practice': '<ellipse cx="12" cy="9.8" rx="2.5" ry="5.3"/>'
                '<ellipse cx="12" cy="9.8" rx="2.5" ry="5.3" transform="rotate(42 12 15)"/>'
                '<ellipse cx="12" cy="9.8" rx="2.5" ry="5.3" transform="rotate(-42 12 15)"/>'
                '<path d="M4.2 15.2c1.3 3.5 4.4 5.2 7.8 5.2s6.5-1.7 7.8-5.2"/>',
    'music':    '<path d="M4.5 10.5v3M8.25 7.5v9M12 9.5v5M15.75 5.5v13M19.5 10v4"/>',
    'scent':    '<path d="M12 3.2c-1.5 1.7-1.5 3.2 0 4.7s1.5 3 0 4.7"/>'
                '<path d="M7.8 16h8.4M9.6 19.5h4.8"/>',
    'food':     '<path d="M4 13h16a8 8 0 0 1-16 0z"/>'
                '<path d="M9.3 9.2c0-1.1 1-1.6 1-2.9M13.7 9.2c0-1.1 1-1.6 1-2.9"/>',
}

def _icon_svg(name: str, color: str, size: int = 14) -> str:
    return (f'<svg viewBox="0 0 24 24" width="{size}" height="{size}" fill="none" '
            f'stroke="{color}" stroke-width="1.7" stroke-linecap="round" '
            f'stroke-linejoin="round">{_ICON[name]}</svg>')

# traditional petal counts per chakra, drawn as line art around the score ring
CHAKRA_PETALS = {'root': 4, 'sacral': 6, 'solar': 10, 'heart': 12,
                 'throat': 16, 'thirdeye': 2, 'crown': 20}

def _score_ring(score: float, color: str, petals: int) -> str:
    C = 55; r = 31; circ = 2 * math.pi * r
    filled = circ * max(0.0, min(100.0, score)) / 100
    offset = 90 if petals == 2 else 0  # third eye: two side petals
    pet = ''.join(
        f'<ellipse cx="{C}" cy="12.5" rx="4.2" ry="7.5" fill="none" stroke="{color}" '
        f'stroke-opacity="0.38" stroke-width="1.3" '
        f'transform="rotate({offset + i * 360 / petals:.1f} {C} {C})"/>'
        for i in range(petals))
    return f'''<svg width="110" height="110" viewBox="0 0 110 110">
      {pet}
      <circle cx="{C}" cy="{C}" r="{r}" fill="none" stroke="#ffffff14" stroke-width="7"/>
      <circle cx="{C}" cy="{C}" r="{r}" fill="none" stroke="{color}" stroke-width="7"
        stroke-linecap="round" stroke-dasharray="{filled:.1f} {circ:.1f}"
        transform="rotate(-90 {C} {C})"/>
      <text x="{C}" y="62" text-anchor="middle" font-size="25" font-weight="900"
        fill="#F4EEFA" font-family="Vazir">{score:.0f}</text>
    </svg>'''

def _week_map(plan: list, active: int) -> str:
    items = ''
    for j, wk in enumerate(plan):
        k = wk['chakra']; c = CHAKRA_COLOR[k]
        extra = f'border-color:{c}88;background:{c}14;' if j == active else ''
        items += f'''<div class="wm-item" style="{extra}">
          <span class="wm-dot" style="background:{c};box-shadow:0 0 7px {c}"></span>
          <span class="wm-w">هفته {_FA_DIGIT[wk['week']]}</span>
          <span class="wm-c">{CHAKRA_LABEL[k]}</span>
          <span class="wm-s">{wk['score']:.0f}</span>
        </div>'''
    return f'<div class="week-map">{items}</div>'

def _pcard(icon: str, title: str, body: str, color: str, extra_cls: str = '') -> str:
    return f'''<div class="pcard {extra_cls}">
      <div class="pcard-h">
        <span class="pcard-n" style="background:{color}1c;border:1px solid {color}55">{_icon_svg(icon, color)}</span>
        <span class="pcard-t">{title}</span>
      </div>{body}</div>'''

def _chips(items: list, color: str) -> str:
    return '<div class="pchips">' + ''.join(
        f'<span class="pchip" style="border-color:{color}44">{x}</span>' for x in items) + '</div>'

def _week_page(plan: list, idx: int, person_block: str, date_str: str) -> str:
    wk = plan[idx]
    k = wk['chakra']; c = CHAKRA_COLOR[k]; score = wk['score']
    t = PROGRAM[k][wk['tier']]
    band_range, band_name = TIER_LABEL[wk['tier']]

    # 1. sleep
    sleep = f'<div class="ln em">{t["sleep"][0]}</div>' + \
            ''.join(f'<div class="ln">{x}</div>' for x in t['sleep'][1:])
    if t.get('sleep_note'):
        sleep += f'<div class="pnote">{t["sleep_note"]}</div>'
    # 2. water + affirmations
    water = f'<div class="ln">{water_note(t)}</div>' + \
            f'<div class="affirm" style="border-color:{c}66">' + \
            ''.join(f'<div class="af-ln">{x}</div>' for x in t['water']) + '</div>'
    # 3. practice
    practice = ''.join(f'<div class="ln{" em" if i == 0 else ""}">{x}</div>'
                       for i, x in enumerate(t['practice']))
    # 4. music
    m0 = t['music'][0].replace(f'فرکانس {CHAKRA_FREQ[k]}', '').strip().lstrip('+ ').strip()
    music_lines = ([m0] if m0 else []) + t['music'][1:]
    music = f'<div class="freq" style="color:{c};border-color:{c}55;background:{c}12">{CHAKRA_FREQ[k]}</div>' + \
            ''.join(f'<div class="ln">{x}</div>' for x in music_lines)
    # 5. scent
    scent = _chips(t['scent'], c)
    if t.get('scent_note'):
        scent += f'<div class="pnote">{t["scent_note"]}</div>'
    # 6. food
    food = f'''<div class="food-cols">
      <div class="food-col">
        <div class="food-h do">بیشتر مصرف شود</div>{_chips(t['food_do'], '#4FB477')}
      </div>
      <div class="food-col">
        <div class="food-h avoid">کمتر مصرف شود</div>{_chips(t['food_avoid'], '#E5484D')}
      </div>
    </div>'''

    return f'''<div class="page">
  <div class="head">
    <div class="brand"><div class="wordmark">شاهراه ثروت</div></div>
    <div class="head-meta">
      {person_block}
      <span>برنامه رشد و تعادل انرژی · هفته {_FA_DIGIT[wk['week']]} از ۴</span><br>
      <span>{html.escape(date_str)}</span>
    </div>
  </div>

  {_week_map(plan, idx)}

  <div class="wk-hero" style="border-color:{c}33">
    <div class="wk-glow" style="background:radial-gradient(60% 90% at 85% 10%, {c}2e 0%, transparent 70%)"></div>
    <div class="wk-body">
      <div class="wk-kicker" style="color:{c}">هفته {_FA_DIGIT[wk['week']]} · چاکرای {CHAKRA_LABEL[k]}</div>
      <div class="wk-title">{CHAKRA_THEME[k]}</div>
      <div class="wk-status">{t['status']}</div>
      <div class="wk-band" style="color:{c};border-color:{c}55;background:{c}14">امتیاز {band_range} · {band_name}</div>
    </div>
    <div class="wk-ring">{_score_ring(score, c, CHAKRA_PETALS[k])}</div>
  </div>

  <div class="pgrid">
    {_pcard('sleep', 'خواب', sleep, c)}
    {_pcard('water', 'آب و بارورسازی آب', water, c)}
    {_pcard('practice', 'تمرین آگاهانه', practice, c)}
    {_pcard('music', 'موسیقی و فرکانس', music, c)}
    {_pcard('scent', 'رایحه و عود', scent, c, 'span2')}
    {_pcard('food', 'الگوی غذایی', food, c, 'span2')}
  </div>

  <div class="foot">
    <div class="disc">این برنامه، یک برنامه رشد و تعادل انرژی است و جایگزین درمان پزشکی، روان‌درمانی یا رژیم‌درمانی نیست. در صورت وجود بیماری یا رژیم خاص، با پزشک خود هماهنگ کنید.</div>
    <div class="foot-brand">شاهراه ثروت</div>
  </div>
</div>'''

def build_html(data: dict, fonts_b64: dict, person_name: str = '', date_str: str = '') -> str:
    # date: explicit arg wins; else the survey's completion date; else today (Iran time)
    date_str = date_str or data.get('report_date') or today_jalali()
    m=data['metrics']
    chak=data['chakras']
    chaklab = CHAKRA_LABEL
    idxlab = INDEX_LABEL
    arche=data['archetype']; arche_fa=data['archetype_fa']
    pos,shadow=ARCHE_DESC.get(arche,('',''))
    dom=data['dominant']; dom_fa=data['dominant_fa']
    overall=data['overall_score'] or 0
    conf=data['confidence']; bal=data['balance'] or 0

    # weakest chakra for the "focus" callout
    weak=min(CHAKRA_ORDER,key=lambda k:(chak.get(k) if chak.get(k) is not None else 999))

    ff=lambda name,w,st='normal':f"@font-face{{font-family:'Vazir';src:url(data:font/ttf;base64,{fonts_b64[name]}) format('truetype');font-weight:{w};font-style:{st};}}"
    fontface=ff('vazir-regular',400)+ff('vazir-medium',500)+ff('vazir-bold',700)+ff('vazir-black',900)

    person_block=f'<span class="meta-name">{html.escape(person_name)}</span>' if person_name else ''

    plan = build_week_plan(chak)
    program_pages = ''.join(_week_page(plan, i, person_block, date_str) for i in range(len(plan)))

    return f'''<!DOCTYPE html><html lang="fa" dir="rtl"><head><meta charset="utf-8">
<style>
{fontface}
*{{margin:0;padding:0;box-sizing:border-box}}
@page{{size:A4;margin:0}}
html,body{{font-family:'Vazir',sans-serif;color:#F4EEFA;background:#0f0b18;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
.page{{width:210mm;height:297mm;padding:14mm 15mm 12mm;background:
   radial-gradient(120% 80% at 80% 0%, #241b38 0%, #160f24 42%, #100b1a 100%);
   position:relative;overflow:hidden;display:flex;flex-direction:column}}
.page::before{{content:'';position:absolute;inset:0;background:
   radial-gradient(50% 40% at 15% 12%, #6a4bd233 0%, transparent 60%);pointer-events:none}}
.page + .page{{page-break-before:always}}

/* header */
.head{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:1px solid #ffffff1f;padding-bottom:14px;position:relative}}
.brand{{display:flex;flex-direction:column;gap:2px}}
.wordmark{{font-weight:900;font-size:30px;letter-spacing:-.5px;line-height:1;
   background:linear-gradient(92deg,#c9b4ff,#8a6bff 60%,#e08bd0);-webkit-background-clip:text;background-clip:text;color:transparent}}
.latin{{font-size:12px;letter-spacing:5px;color:#8f83aa;text-transform:uppercase;font-weight:500;margin-top:5px}}
.tagline{{font-size:12px;color:#a99dc4;margin-top:6px}}
.head-meta{{text-align:left;font-size:11.5px;color:#8f83aa;line-height:1.9}}
.meta-name{{display:block;color:#e6dcf7;font-size:14px;font-weight:700}}

/* section scaffolding */
.sec{{margin-top:20px}}
.sec-h{{display:flex;align-items:baseline;gap:10px;margin-bottom:12px}}
.sec-n{{font-size:12px;color:#7c6fa0;font-weight:700;letter-spacing:1px}}
.sec-t{{font-size:17px;font-weight:700;color:#efe8fb}}
.sec-line{{flex:1;height:1px;background:#ffffff14;align-self:center;margin-inline-start:6px}}

/* hero: overall + spine */
.hero{{display:grid;grid-template-columns:1.05fr 1fr;gap:24px;margin-top:18px;align-items:center}}
.overall-card{{background:#ffffff08;border:1px solid #ffffff14;border-radius:18px;padding:22px 24px;position:relative}}
.ov-k{{font-size:12.5px;color:#a99dc4}}
.ov-num{{font-size:64px;font-weight:900;line-height:1;margin:6px 0 2px;
   background:linear-gradient(180deg,#d9c8ff,#8f6bff);-webkit-background-clip:text;background-clip:text;color:transparent}}
.ov-sub{{font-size:12.5px;color:#9a8fb0}}
.chips{{display:flex;gap:8px;margin-top:16px;flex-wrap:wrap}}
.chip{{font-size:11px;padding:5px 11px;border-radius:20px;background:#ffffff0f;border:1px solid #ffffff1a;color:#c7bce0}}
.chip b{{color:#efe8fb;font-weight:700}}

.spine{{display:flex;flex-direction:column;gap:9px}}
.spine-row{{display:grid;grid-template-columns:26px 1fr 10px 52px;align-items:center;gap:9px}}
.spine-num{{font-size:12px;color:#cabfe0;text-align:center;font-weight:700}}
.spine-track{{height:7px;background:#ffffff10;border-radius:6px;overflow:hidden;direction:ltr}}
.spine-fill{{height:100%;border-radius:6px}}
.spine-dot{{width:9px;height:9px;border-radius:50%}}
.spine-label{{font-size:12.5px;color:#cabfe0}}

/* archetype block */
.arche{{display:grid;grid-template-columns:auto 1fr;gap:18px;background:
   linear-gradient(100deg,#2a1f45 0%,#1b1430 100%);border:1px solid #ffffff14;border-radius:18px;padding:20px 22px;align-items:center}}
.arche-badge{{width:92px;height:92px;border-radius:50%;display:flex;flex-direction:column;align-items:center;justify-content:center;
   background:radial-gradient(circle at 50% 35%,#6a4bd2,#3a2a6e);border:1px solid #ffffff26;box-shadow:0 0 24px #6a4bd255}}
.arche-badge .fa{{font-size:18px;font-weight:900;color:#fff}}
.arche-badge .en{{font-size:9px;letter-spacing:2px;color:#d9ccff;margin-top:2px}}
.arche-body h3{{font-size:15px;color:#efe8fb;margin-bottom:8px}}
.arche-body .row{{font-size:12.5px;color:#b7abd0;margin-bottom:5px;line-height:1.7}}
.arche-body .row b{{color:#e6dcf7}}
.arche-body .k{{color:#8f83aa}}

/* radar + interpretation two-col */
.two{{display:grid;grid-template-columns:340px 1fr;gap:22px;align-items:start}}
.radar-wrap{{background:#ffffff07;border:1px solid #ffffff12;border-radius:18px;padding:12px;display:flex;justify-content:center}}
.legend{{display:flex;flex-direction:column;gap:9px}}

/* bars */
.bar-row{{display:grid;grid-template-columns:74px 1fr 30px 92px;align-items:center;gap:10px;margin-bottom:11px}}
.bar-label{{font-size:12.5px;color:#cabfe0}}
.bar-track{{height:8px;background:#ffffff10;border-radius:6px;overflow:hidden;direction:ltr}}
.bar-fill{{height:100%;border-radius:6px;background:linear-gradient(90deg,#7c5cff,#b46fd0)}}
.bar-val{{font-size:12px;color:#e6dcf7;font-weight:700;text-align:center}}
.bar-tag{{font-size:10.5px;padding:3px 8px;border-radius:12px;text-align:center;white-space:nowrap}}
.tag-strength{{background:#4fb47722;color:#7fe0a6;border:1px solid #4fb47740}}
.tag-moderate{{background:#3da5d922;color:#7fcdf0;border:1px solid #3da5d940}}
.tag-low{{background:#f5c54222;color:#f0d47f;border:1px solid #f5c54240}}
.tag-needs_attention{{background:#e5484d22;color:#f09a9d;border:1px solid #e5484d40}}

/* callouts */
.callouts{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:6px}}
.call{{background:#ffffff08;border:1px solid #ffffff14;border-radius:14px;padding:15px 16px}}
.call .k{{font-size:11px;color:#8f83aa;margin-bottom:6px}}
.call .v{{font-size:14px;color:#efe8fb;font-weight:700;margin-bottom:7px}}
.call .d{{font-size:12px;color:#b0a4c9;line-height:1.75}}

/* interpretation rows */
.interp-row{{border-inline-start:2px solid;padding:8px 12px;margin-bottom:9px;background:#ffffff06;border-radius:0 10px 10px 0}}
.interp-row .t{{font-size:12.5px;color:#e6dcf7;font-weight:700;display:flex;justify-content:space-between}}
.interp-row .s{{font-size:11.5px;color:#a99dc4;margin-top:3px;line-height:1.7}}

/* ===== 4-week program pages ===== */
.week-map{{display:grid;grid-template-columns:repeat(4,1fr);gap:9px;margin-top:14px}}
.wm-item{{display:flex;align-items:center;gap:7px;padding:8px 11px;border-radius:12px;
   background:#ffffff07;border:1px solid #ffffff14;font-size:11px;color:#a99dc4}}
.wm-dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
.wm-w{{font-weight:700;color:#e6dcf7;white-space:nowrap}}
.wm-c{{white-space:nowrap}}
.wm-s{{margin-inline-start:auto;font-weight:700;color:#cabfe0}}

.wk-hero{{position:relative;overflow:hidden;display:grid;grid-template-columns:1fr auto;gap:18px;align-items:center;
   background:linear-gradient(100deg,#241b3e 0%,#181128 100%);border:1px solid #ffffff14;border-radius:18px;
   padding:18px 22px;margin-top:14px}}
.wk-glow{{position:absolute;inset:0;pointer-events:none}}
.wk-body{{position:relative}}
.wk-kicker{{font-size:12px;font-weight:700;letter-spacing:.3px}}
.wk-title{{font-size:21px;font-weight:900;color:#efe8fb;margin:5px 0 7px}}
.wk-status{{font-size:11.5px;color:#b7abd0;line-height:1.85;max-width:56ch}}
.wk-band{{display:inline-block;font-size:11px;font-weight:700;padding:5px 12px;border-radius:20px;
   border:1px solid;margin-top:10px}}
.wk-ring{{position:relative;display:flex;align-items:center;justify-content:center}}

.pgrid{{display:grid;grid-template-columns:1fr 1fr;gap:11px;margin-top:14px;flex:1;align-content:stretch}}
.pcard{{background:#ffffff07;border:1px solid #ffffff12;border-radius:14px;padding:13px 15px}}
.pcard.span2{{grid-column:1 / -1}}
.pcard-h{{display:flex;align-items:center;gap:9px;margin-bottom:9px}}
.pcard-n{{width:24px;height:24px;border-radius:7px;display:flex;align-items:center;justify-content:center;flex-shrink:0}}
.pcard-n svg{{display:block}}
.pcard-t{{font-size:13px;font-weight:700;color:#efe8fb}}
.ln{{font-size:11px;color:#b7abd0;line-height:1.85}}
.ln.em{{color:#e6dcf7;font-weight:700}}
.pnote{{font-size:10px;color:#8f83aa;line-height:1.75;margin-top:7px;padding-top:7px;border-top:1px dashed #ffffff14}}
.affirm{{border-inline-start:2px solid;padding:7px 11px;margin-top:8px;background:#ffffff05;border-radius:0 10px 10px 0}}
.af-ln{{font-size:11px;color:#d9cdf0;line-height:1.95;font-weight:500}}
.freq{{display:inline-block;font-size:15px;font-weight:900;letter-spacing:.5px;padding:4px 14px;
   border-radius:10px;border:1px solid;margin-bottom:8px;direction:ltr}}
.pchips{{display:flex;flex-wrap:wrap;gap:6px}}
.pchip{{font-size:10.5px;padding:4px 10px;border-radius:14px;background:#ffffff0a;
   border:1px solid;color:#cabfe0;white-space:nowrap}}
.food-cols{{display:grid;grid-template-columns:1.4fr 1fr;gap:14px}}
.food-h{{font-size:11px;font-weight:700;margin-bottom:7px}}
.food-h.do{{color:#7fe0a6}}
.food-h.avoid{{color:#f09a9d}}

/* footer / disclaimer */
.foot{{margin-top:auto;border-top:1px solid #ffffff14;padding-top:11px;display:flex;justify-content:space-between;align-items:center}}
.disc{{font-size:10px;color:#7c7196;line-height:1.7;max-width:70%}}
.foot-brand{{font-size:11px;color:#9a8fb0;font-weight:700}}
</style></head>
<body>
<!-- ===================== PAGE 1 ===================== -->
<div class="page">
  <div class="head">
    <div class="brand">
      <div class="wordmark">شاهراه ثروت</div>
      <div class="tagline">تحلیل انرژی و تعادل درونی</div>
    </div>
    <div class="head-meta">
      {person_block}
      <span>گزارش تحلیل فردی</span><br>
      <span>{html.escape(date_str)}</span>
    </div>
  </div>

  <div class="hero">
    <div class="overall-card">
      <div class="ov-k">امتیاز کلی تعادل</div>
      <div class="ov-num">{overall:.0f}<span style="font-size:24px">٪</span></div>
      <div class="ov-sub">میانگین هفت مرکز انرژی شما</div>
      <div class="chips">
        <div class="chip">تعادل کلی <b>{bal:.0f}</b></div>
        <div class="chip">اطمینان داده <b>{conf:.0f}٪</b></div>
        <div class="chip">مرکز غالب <b>{dom_fa}</b></div>
      </div>
    </div>
    {_spine(chak)}
  </div>

  <div class="sec">
    <div class="sec-h"><span class="sec-n">۰۱</span><span class="sec-t">آرکتایپ غالب شما</span><span class="sec-line"></span></div>
    <div class="arche">
      <div class="arche-badge"><div class="fa">{arche_fa}</div><div class="en">{arche.upper()}</div></div>
      <div class="arche-body">
        <h3>الگوی انرژی غالب شما بر پایه مرکز «{dom_fa}» شکل گرفته است</h3>
        <div class="row"><span class="k">نقش مثبت:</span> <b>{pos}</b></div>
        <div class="row"><span class="k">سایه احتمالی:</span> <b>{shadow}</b></div>
      </div>
    </div>
  </div>

  <div class="sec">
    <div class="sec-h"><span class="sec-n">۰۲</span><span class="sec-t">نمای هفت مرکز انرژی</span><span class="sec-line"></span></div>
    <div class="two">
      <div class="radar-wrap">{_radar_svg(chak)}</div>
      <div class="legend">{_bars(m, ['crown','thirdeye','throat','heart','solar','sacral','root'], chaklab)}</div>
    </div>
  </div>

  <div class="foot">
    <div class="disc">این گزارش ابزار خودشناسی و توسعه فردی است و جایگزین تشخیص پزشکی، روان‌پزشکی یا مشاوره مالی نیست. وزن‌ها و امتیازها بر پایه نسخه اولیه محصول محاسبه شده‌اند.</div>
    <div class="foot-brand">شاهراه ثروت</div>
  </div>
</div>

<!-- ===================== PAGE 2 ===================== -->
<div class="page">
  <div class="head">
    <div class="brand"><div class="wordmark">شاهراه ثروت</div></div>
    <div class="head-meta"><span>ادامه گزارش · ابعاد و مسیر رشد</span></div>
  </div>

  <div class="sec">
    <div class="sec-h"><span class="sec-n">۰۳</span><span class="sec-t">شاخص‌های تکمیلی زندگی</span><span class="sec-line"></span></div>
    <div class="legend">{_bars(m, INDEX_ORDER, idxlab)}</div>
  </div>

  <div class="sec">
    <div class="sec-h"><span class="sec-n">۰۴</span><span class="sec-t">تحلیل کلیدی</span><span class="sec-line"></span></div>
    <div class="callouts">
      <div class="call">
        <div class="k">نقطه قوت غالب</div>
        <div class="v">{dom_fa} · {chak[dom]:.0f}٪</div>
        <div class="d">این مرکز به عنوان منبع اصلی قدرت شما دیده می‌شود. از آن آگاهانه برای حمایت از حوزه‌های ضعیف‌تر استفاده کنید.</div>
      </div>
      <div class="call">
        <div class="k">حوزه نیازمند تمرکز</div>
        <div class="v">{chaklab[weak]} · {chak[weak]:.0f}٪</div>
        <div class="d">{BAND_SUGGEST[m[weak]['level']]}</div>
      </div>
    </div>
  </div>

  <div class="sec">
    <div class="sec-h"><span class="sec-n">۰۵</span><span class="sec-t">تفسیر و پیشنهاد هر مرکز</span><span class="sec-line"></span></div>
    <div>{_interp_rows(m, chaklab)}</div>
  </div>

  <div class="foot">
    <div class="disc">نتایج بر اساس پاسخ‌های ثبت‌شده در پرسشنامه ۷۰ سؤالی تولید شده‌اند. برای اعتبار بیشتر، پاسخ‌ها را کامل و صادقانه تکمیل کنید.</div>
    <div class="foot-brand">شاهراه ثروت</div>
  </div>
</div>

<!-- ============== 4-WEEK PROGRAM PAGES ============== -->
{program_pages}
</body></html>'''



# ===== renderer =====
def _render_pages(browser, jobs) -> None:
    pg = browser.new_page()
    for html_str, pdf_path in jobs:
        pg.set_content(html_str, wait_until='networkidle')
        pg.pdf(path=pdf_path, format='A4', print_background=True,
               margin={'top':'0','bottom':'0','left':'0','right':'0'})
    pg.close()

def render_html_to_pdf(html_str: str, pdf_path: str) -> None:
    with sync_playwright() as p:
        b = p.chromium.launch(args=['--no-sandbox'])
        _render_pages(b, [(html_str, pdf_path)])
        b.close()

def safe_pdf_name(name: str) -> str:
    """'<sanitized person>.pdf' — Persian letters kept, filesystem junk not."""
    safe = re.sub(r'[^\w\u0600-\u06FF\- ]', '', name).strip() if name else ''
    safe = re.sub(r'\s+', '_', safe)
    return f"{safe}.pdf" if safe else "chakra-report.pdf"

def render_reports(datas: list, out_dir: str, name_override: str = '',
                   date_str: str = '') -> list:
    """Render one PDF per scored respondent into out_dir, reusing a single
    Chromium instance (a per-PDF launch is ~2s of pure overhead — fatal for a
    50-person batch inside an HTTP timeout). Files are named after the person,
    deduped with _2, _3… on collisions. name_override only applies when there
    is exactly one respondent — a batch reads every name from the file.
    Returns [(pdf_path, person_name, data), ...] in file order."""
    fonts = _load_fonts()
    jobs, meta, used = [], [], set()
    for i, data in enumerate(datas, 1):
        person = (name_override or '').strip() if len(datas) == 1 else ''
        person = person or data.get('person_name', '') or f'respondent-{i}'
        fname = safe_pdf_name(person)
        stem, n = fname[:-4], 2
        while fname in used:
            fname, n = f'{stem}_{n}.pdf', n + 1
        used.add(fname)
        pdf_path = os.path.join(out_dir, fname)
        jobs.append((build_html(data, fonts, person_name=person,
                                date_str=date_str), pdf_path))
        meta.append((pdf_path, person, data))
    with sync_playwright() as p:
        b = p.chromium.launch(args=['--no-sandbox'])
        _render_pages(b, jobs)
        b.close()
    return meta

# ===== main =====
if __name__ == '__main__':
    in_xlsx = sys.argv[1]
    out_pdf = sys.argv[2]
    person  = sys.argv[3] if len(sys.argv) > 3 else ''
    datestr = sys.argv[4] if len(sys.argv) > 4 else ''
    datas = score_workbook_all(in_xlsx)
    if len(datas) > 1:
        # batch file: the output argument is a directory, one PDF per person
        os.makedirs(out_pdf, exist_ok=True)
        reports = render_reports(datas, out_pdf, date_str=datestr)
        print(json.dumps({'ok': True, 'count': len(reports),
                          'reports': [{'pdf': p, 'person': n,
                                       'overall': d['overall_score'],
                                       'confidence': d['confidence']}
                                      for p, n, d in reports]},
                         ensure_ascii=False, indent=1))
        sys.exit(0)
    data  = datas[0]
    fonts = _load_fonts()
    html_str = build_html(data, fonts, person_name=person, date_str=datestr)
    render_html_to_pdf(html_str, out_pdf)
    print(json.dumps({'ok': True, 'pdf': out_pdf,
                      'overall': data['overall_score'],
                      'dominant': data['dominant_fa'],
                      'archetype': data['archetype'],
                      'confidence': data['confidence'],
                      'weeks': build_week_plan(data['chakras'])}, ensure_ascii=False))